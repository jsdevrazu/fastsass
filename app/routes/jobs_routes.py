from fastapi import APIRouter, Depends, Query
from pymongo.errors import PyMongoError
from app.utils.error_handler import api_error
from app.validation_schema.post_validation_schema import JobSchema
from bson import ObjectId
from app.dependencies.subscription import require_active_subscription
from app.auth.jwt_handler import require_role, get_current_user, optional_get_current_user
from app.utils.helper import slugify, calculate_match_percentage, serialize_notification
from datetime import datetime, timedelta, timezone
from app.db.mongo import db
from app.validation_schema.post_validation_schema import UpdateJobSchema, ApplicationStatus, ApplicationSchema, ReviewsRatinsSchema, CoverLetterInput
from datetime import datetime
from fastapi.responses import StreamingResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from typing import Optional
import os
from PyPDF2 import PdfReader
from openai import OpenAI
from app.configs.settings import settings
from app.constant.index import AI_MODEL
from app.socket import emit_event


client = OpenAI(
  base_url="https://openrouter.ai/api/v1",
  api_key=settings.open_api_key,
)


def extract_structured_data_from_resume(text: str) -> str:
    prompt = f"""You are a professional resume parser. Extract key structured information from the following resume text:
    Return the result in JSON format with the following fields:
    - full_name
    - title
    - summary or about us
    - email
    - phone
    - location
    - linkedin
    - website
    - github
    - education (list of degree, institution, year, gpa, year: Graduated Year)
    - experience (list of job_title, company, duration, description: list of strings, location: default value remote if not found location)
    - skills (list of skills)

    Respond only with JSON format and no explanation.

    Resume Content:
    \"\"\"
    {text}
    \"\"\"
    """
    response = client.chat.completions.create(
        model=AI_MODEL,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2,
    )

    raw_output = response.choices[0].message.content or ''

    try:
      return raw_output
    except Exception as error:
        return ''
    
def extract_text_from_pdf(file_path: str) -> str:
    full_path = os.path.join(os.getcwd(), file_path)
    if not os.path.exists(full_path):
        raise FileNotFoundError(f"Resume not found at {full_path}")

    reader = PdfReader(full_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text() or ""
    return text.strip()



router = APIRouter()


@router.post('/post')
def post_job(
    body:JobSchema,
    user=Depends(require_active_subscription)
):
    try:
        now = datetime.now()
        payload = {
            **body.model_dump(),
            "slug": slugify(body.title),
            "created_at": now,
            "updated_at": now,
            "post_by": ObjectId(user['_id']),
            "application_dead_line": now + timedelta(days=30),
            "company_name": ObjectId(user['company_id'])
        }

        db.jobs.insert_one(payload)

        db.users.find_one_and_update(
            {"email": user['email'],},
            {"$inc": {"feature.used_job": 1}}
        )

        return {
            "message": "Job post successfully"
        }
    except PyMongoError:
        api_error(500, "Invalid Database Request")


@router.get("/jobs")
def get_all_jobs(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1),
    title: str = Query('', alias="title"),
    location: str = Query('', alias="location"),
    job_type: str = Query('', alias="job_type"),
    min_salary: int = Query(None, alias="min_salary"),
    max_salary: int = Query(None, alias="max_salary"),

): 
    skip = (page - 1) * limit
    try:
        filters = {}

        if title:
            filters["title"] = {"$regex": title, "$options": "i"}
        
        if location:
            filters["location"] = {"$regex": location, "$options": "i"}
        
        if job_type:
            filters["job_type"] = job_type
        
        if min_salary is not None:
            filters["min_salary"] = {"$gte": min_salary}
        
        if max_salary is not None:
            filters["max_salary"] = {"$lte": max_salary}

        filters["job_status"] = "active"

        pipeline = [
            {"$match": filters},
            {
                "$lookup": {
                    "from": "companies", 
                    "localField": "company_name",
                    "foreignField": "_id", 
                    "as": "company"
                }
            },
            {"$unwind": {"path": "$company", "preserveNullAndEmptyArrays": True}}, 
            {"$skip": skip},
            {"$limit": limit},
            {
                "$project": {
                    "_id": {"$toString": "$_id"}, 
                    "title": 1,
                    "min_salary": 1,
                    "max_salary": 1,
                    "job_type": 1,
                    "slug": 1,
                    "created_at": 1,
                    "updated_at": 1,
                    "location": 1,
                    "body": 1,
                    "job_type": 1,
                    "skills": 1,
                    "questions": 1,
                    "apply_settings": 1,
                    "job_status": 1,
                    "company._id": {"$toString": "$company._id"},
                    "company.name": 1,
                    "company.logo": 1
                }
            }
        ]

        cursor = list(db.jobs.aggregate(pipeline))

        total = db.jobs.count_documents(filters)

        return {
            "message": "success",
            "total": total,
            "page": page,
            "limit": limit,
            "jobs": cursor
        }

    except PyMongoError as error:
        api_error(500, str(error))


@router.get('/{slug}/edit/detail')
def get_edit_job_detail(
    slug: str,
    user = Depends(require_role('employer'))
):
    job = db.jobs.find_one({"slug": slug, "post_by": ObjectId(user["_id"])})

    if not job:
        api_error(404, 'Job Not Found')
    
    job['_id'] = str(job['_id'])
    job['post_by'] = str(job['post_by'])
    job['company_name'] = str(job['company_name'])

    return {
        "message": "success",
        "job": job
    } 


@router.get('/{slug}/detail')
def get_single_job(
    slug: str,
    user: Optional[dict] = Depends(optional_get_current_user)
):
    pipeline = [
        {"$match": {"slug": slug}},
        {
            "$lookup": {
                "from": "companies",
                "localField": "company_name",
                "foreignField": "_id",
                "as": "company"
            }
        },
        {"$unwind": {"path": "$company", "preserveNullAndEmptyArrays": True}},
    ]

    if user:
        pipeline += [
            {
                "$lookup": {
                    "from": "save_jobs",
                    "let": {"jobId": "$_id"},
                    "pipeline": [
                        {
                            "$match": {
                                "$expr": {
                                    "$and": [
                                        {"$eq": ["$job_id", "$$jobId"]},
                                        {"$eq": ["$user_id", ObjectId(user["_id"])]}
                                    ]
                                }
                            }
                        },
                        {"$limit": 1}
                    ],
                    "as": "saved"
                }
            },
            {
                "$addFields": {
                    "isBookmarked": {"$gt": [{"$size": "$saved"}, 0]}
                }
            }
        ]
    else:
        pipeline += [
            {
                "$addFields": {
                    "isBookmarked": False
                }
            }
        ]

    pipeline += [
        {
            "$project": {
                "_id": {"$toString": "$_id"},
                "title": 1,
                "body": 1,
                "min_salary": 1,
                "max_salary": 1,
                "location": 1,
                "job_type": 1,
                "slug": 1,
                "created_at": 1,
                "updated_at": 1,
                "skills": 1,
                "experience_level": 1,
                "questions": 1,
                "apply_settings": 1,
                "application_email": 1,
                "application_dead_line": 1,
                "application_link": 1,
                "job_status": 1,
                "isBookmarked": 1,
                "company._id": {"$toString": "$company._id"},
                "company.name": 1,
                "company.logo": 1,
                "company.company_description": 1,
            }
        }
    ]

    cursor_job = db.jobs.aggregate(pipeline)
    job_list = list(cursor_job)
    if not job_list:
        api_error(404, "Job not found")

    return {
        "message": "success",
        "job": job_list[0]
    } 

@router.patch("/update/{job_id}")
def update_job(
    job_id: str,
    body: UpdateJobSchema,
    user=Depends(require_active_subscription)
):
    try:

        filtered_data = {k: v for k, v in body.model_dump().items() if v is not None}

        if "title" in filtered_data:
            filtered_data['slug'] = slugify(filtered_data['title'])

        if not filtered_data:
            api_error(400, "No valid fields to update")

        job = db.jobs.find_one_and_update(
            {"_id": ObjectId(job_id), "post_by": ObjectId(user['_id'])},
            {"$set": filtered_data},
            return_document=True
        )

        if not job:
            api_error(404, "Job not found")

        return {
            "message": "Update Successfully"
        }
    except PyMongoError as error:
        api_error(500, str(error))
    

@router.delete("/delete/{job_id}")
def delete_job(job_id: str, user=Depends(require_active_subscription)):
    try:
        job = db.jobs.delete_one({"_id": ObjectId(job_id), "post_by": ObjectId(user['_id'])})

        if job.deleted_count == 0:
            api_error(400, "Job not found")
        
        return {
            "message": "Job Delete Successfully"
        }
    except PyMongoError as error:
        api_error(500, str(error))


@router.post('/apply/{job_id}')
async def apply_job(
    body: ApplicationSchema,
    job_id: str,
    user=Depends(require_role('job_seeker'))
):
    try:
        job = db.jobs.find_one({"_id": ObjectId(job_id)})
        application = db.applications.find_one({
            "job_id": ObjectId(job_id),
            "applicate": {"$in": [ObjectId(user['_id'])]}
        })


        if not job:
            api_error(404, "Job not found")
        
        if application:
            api_error(400, 'You have already apply this job')
        
        payload = {
            **body.model_dump(),
            "job_id": ObjectId(job_id),
            "applicate": ObjectId(user['_id']),
            "created_at": datetime.now(),
            "updated_at": datetime.now(),
            "application_status": 'Pending'
        }

        user_id = str(job['post_by'])
        applicant_name = f"{user.get("first_name")} {user.get("last_name")}"

        db.applications.insert_one(payload)

        await emit_event(str(job['post_by']), 'NEW_APPLICATION', {"application": 1})
        

        payload = {
        "user_id": ObjectId(user_id),
        "type": "MENTION",
        "status": "UNREAD",
        "title": f"New Job Application Received!",
        "message": f"{applicant_name} has just applied for your job post: \"{job.get("title")}\". Review the application to take the next step.",
        "created_at": datetime.now(),
        "updated_at": datetime.now(),
        }
        new_notification = db.notifications.insert_one(payload)
        notification = {
            **payload,
            "_id": str(new_notification.inserted_id),
            "user_id": str(user_id)
        }

        await emit_event(user_id, "NEW_APPLICATION_SUBMIT", serialize_notification(notification))


        return {
            "message": "Job Apply Successfully"
        }
    
    except PyMongoError as error:
        api_error(500, str(error))

@router.get('/get-apply-jobs')
def get_apply_jobs(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1),
    user=Depends(require_role('job_seeker'))
):
    skip = (page - 1) * limit
    try:
        pipeine = [
            {"$match": {
                "applicate": ObjectId(user['_id'])
            }},
            {
            "$lookup": {
                "from": "jobs",            
                "localField": "job_id",       
                "foreignField": "_id",    
                "as": "job"
            }
            },
            {"$unwind": {"path": "$job", "preserveNullAndEmptyArrays": True}},
            {
                "$lookup": {
                    "from": "companies",
                    "localField": "job.company_name",
                    "foreignField": "_id",
                    "as": "company"
                }
            },
            {"$unwind": {"path": "$company", "preserveNullAndEmptyArrays": True}},
            {"$skip": skip},
            {"$limit": limit},
            {
                "$project": {
                    "_id": {"$toString": "$_id"},
                    "job_id": {"$toString": "$job._id"},
                    "application_status": 1,
                    "job.title": 1,
                    "job.created_at": 1,
                    "job.location": 1,
                    "job.slug": 1,
                    "company_name": "$company.name"

                }
            }
        ]

        applications = list(db.applications.aggregate(pipeine))
        total = db.applications.count_documents({"applicate": ObjectId(user['_id'])})
        

        return {
            "message": "success",
            "page": page,
            "total": total,
            "limit": limit,
            "applications": applications

        }

    except PyMongoError as error:
        api_error(500, str(error))

@router.get("/recent-applied")
def recent_applied_jobs(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1),
    user=Depends(require_role("job_seeker"))
):
        skip = (page - 1) * limit
        ten_days_ago = datetime.now(timezone.utc) - timedelta(days=10)

        pipeline = [
            {
                "$match": {
                    "applicate": ObjectId(user["_id"]),
                    "created_at": {"$gte": ten_days_ago}
                }
            },
            {
                "$lookup": {
                    "from": "jobs",
                    "localField": "job_id",
                    "foreignField": "_id",
                    "as": "job"
                }
            },
            {"$unwind": {"path": "$job", "preserveNullAndEmptyArrays": True}},
            {
                "$lookup": {
                    "from": "companies",
                    "localField": "job.company_name",
                    "foreignField": "_id",
                    "as": "company"
                }
            },
            {"$unwind": {"path": "$company", "preserveNullAndEmptyArrays": True}},
            {"$skip": skip},
            {"$limit": limit},
            {
                "$project": {
                    "_id": {"$toString": "$_id"},
                    "job_id": {"$toString": "$job._id"},
                    "application_status": 1,
                    "job.title": 1,
                    "job.created_at": 1,
                    "job.location": 1,
                    "job.slug": 1,
                    "company_name": "$company.name"

                }
            },
            {"$sort": {"created_at": -1}}
        ]

        applications = list(db.applications.aggregate(pipeline))
        total = db.applications.count_documents({"applicate": ObjectId(user['_id'])})

        return {
            "message": "success",
            "page": page,
            "total": total,
            "limit": limit,
            "applications": applications
        }


@router.get('/recommended')
def get_recommended_jobs(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1),
    user=Depends(get_current_user)
):
    skip = (page - 1) * limit
    preferences = user.get("preferences_feature", {})
    job_type = preferences.get("job_type")
    work_location = preferences.get("work_location")
    salary_range = preferences.get("salary_range")
    experience_level = preferences.get("experience_level")
    preferred_industry = preferences.get("preferred_industry")

    filters = {}

    if job_type:
        filters["job_type"] = job_type

    if work_location:
        filters["location"] = work_location

    if experience_level:
        filters["experience_level"] = experience_level

    if preferred_industry:
        filters["industry"] = preferred_industry

    if salary_range:
        try:
            min_str, max_str = salary_range.lower().replace("k", "000").split("-")
            preferred_min = int(min_str.strip())
            preferred_max = int(max_str.strip())

            filters["$and"] = [
                {"min_salary": {"$lte": preferred_max}},
                {"max_salary": {"$gte": preferred_min}},
            ]
        except Exception:
            pass   
        
    applied_jobs = list(db.applications.find(
            {"applicate": ObjectId(user['_id'])},
            {"job_id": 1}
        ))
    
    applied_job_ids = [item["job_id"] for item in applied_jobs]

    filters["_id"] = {"$nin": applied_job_ids}


    pipeline = [
            {
                "$match": filters
            },
            {
                "$lookup": {
                    "from": "companies",
                    "localField": "company_name",
                    "foreignField": "_id",
                    "as": "company"
                }
            },
            {
                "$unwind": {
                    "path": "$company",
                    "preserveNullAndEmptyArrays": True
                }
            },
            {
                "$project": {
                    "title": 1,
                    "job_type": 1,
                    "location": 1,
                    "slug": 1,
                    "post_by": 1,
                    "company_name": "$company.name",
                    "created_at": 1
                }
            },
            {"$skip": skip},
            {"$limit": limit}
        ]

    recommended_jobs = list(db.jobs.aggregate(pipeline))

    total = db.jobs.count_documents({
            **filters,
            "_id": {"$nin": applied_job_ids}
    })

    for job in recommended_jobs:
            job['_id'] = str(job['_id'])
            job['post_by'] = str(job['post_by'])

    return {
            "message": "success",
            "total": total,
            "page": page,
            "limit": limit,
            "recommended_jobs": recommended_jobs
        }


@router.post('/save/{job_id}')
def save_job(job_id: str, user=Depends(require_role('job_seeker'))):
        job = db.jobs.find_one({"_id": ObjectId(job_id)})
        if not job:
            api_error(404, "Job not found")

        save_job = db.save_jobs.find_one({
             "job_id": ObjectId(job_id),
             "user_id": {"$in": [ObjectId(user['_id'])]}
        })

        if save_job:
            api_error(400, "You already saved this job")


        db.save_jobs.insert_one({
            "job_id": ObjectId(job_id),
            "user_id": ObjectId(user['_id']),
            "created_at": datetime.now(),
            "updated_at": datetime.now()
        })

        return {
            "message": "Job Saved Successfully"
        }


@router.get("/stats")
def dashboard_stats(user=Depends(require_role('job_seeker'))):
    application = db.applications.count_documents({"applicate": ObjectId(user['_id'])})
    interviewed = db.applications.count_documents({"application_status": 'Interviewed'})
    save_jobs = db.save_jobs.count_documents({"user_id": ObjectId(user['_id'])})
    total_views_pipeline = [
        {
            "$group": {
                "_id": ObjectId(user['_id']),
                "total_views": {"$sum": "$profile_view"}
            }
        }
    ]

    result = list(db.users.aggregate(total_views_pipeline))
    total_views = result[0]['total_views'] if result else 0

    return {
        "message": "success",
        "application": application,
        "interviewed": interviewed,
        "save_jobs": save_jobs,
        "profile_views": total_views

    }

@router.get('/saved-jobs')
def get_saved_jobs(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1),
    user=Depends(require_role("job_seeker"))
):
    skip = (page-1) * limit
    try:
        
        pipeline = [
            {"$match": {"user_id": ObjectId(user['_id'])}},
            {
                "$lookup":{
                    "from": "jobs",
                    "localField": "job_id",
                    "foreignField": "_id", 
                    "as": "saved_jobs"

                }
            },
            {"$unwind": {"path": "$saved_jobs", "preserveNullAndEmptyArrays": True}}, 
            {
                "$lookup": {
                    "from": "companies",
                    "localField": "saved_jobs.company_name",
                    "foreignField": "_id",
                    "as": "company"
                }
            },
            {"$unwind": {"path": "$company", "preserveNullAndEmptyArrays": True}},
            {"$skip": skip},
            {"$limit": limit},
            {
                "$project": {
                   "_id": {"$toString": "$_id"}, 
                   "user_id": {"$toString": "$user_id"}, 
                   "saved_jobs._id": {"$toString": "$saved_jobs._id"}, 
                    "saved_jobs.title": 1,
                    "saved_jobs.location": 1,
                    "saved_jobs.job_type": 1,
                    "company_name": "$company.name",
                    "saved_jobs.body": 1,
                    "saved_jobs.slug": 1,
                    "saved_jobs.created_at": 1,
                    "saved_jobs.updated_at": 1,
                }
            }
        ]

        total = db.save_jobs.count_documents({})
        jobs = list(db.save_jobs.aggregate(pipeline))


        return {
            "message": "success",
            "page": page,
            "limit": limit,
            "total": total,
            "jobs": jobs
        }
    except PyMongoError as error:
        api_error(500, str(error))

@router.post('/status/{application_id}')
async def update_job_status(application_id: str, body: ApplicationStatus):
    job = db.applications.find_one_and_update(
        {"_id": ObjectId(application_id)},
        {"$set": {"application_status": body.application_status}},
        return_document=True
    )

    user_id = str(job.get("applicate", ""))
    
    if not job:
        api_error(404, 'Job not found')
    if body.application_status == 'Interviewed':
        await emit_event(str(job.get("applicate", "")), "APPLICATION_STATUS", {"interviewed": 1})
    
    payload = {
        "user_id": ObjectId(user_id),
        "type": "RESOURCE_UPDATE",
        "status": "UNREAD",
        "title": f"Your job has been {body.application_status}!",
        "message": f"`Congratulations! Your application has been {body.application_status} state",
        "created_at": datetime.now(),
        "updated_at": datetime.now(),
    }
    new_notification = db.notifications.insert_one(payload)
    notification = {
        **payload,
        "_id": str(new_notification.inserted_id),
        "user_id": str(user_id)
    }

    await emit_event(user_id, "NOTIFICATION", serialize_notification(notification))

    return {
        "message": "Job update status success",
    }

@router.get('/update-view/{slug}')
async def update_job_view(slug:str):

    job = db.jobs.find_one_and_update(
        {"slug": slug},
        {"$inc": {"job_view": 1}}
    )

    if not job:
        api_error(404, "Jon not found")
    
    await emit_event(str(job['post_by']), "JOB_VIEW_UPDATE", {"job_view": 1})

    return {
        "message": "View Update"
    }


@router.get('/applications/export')
def export_applications(user=Depends(require_role('employer'))):
    try:
        applications_cursor = db.applications.find(
            {"applicate": ObjectId(user['_id'])},
            {"job_id": 1, "application_status": 1, "created_at": 1}
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Job Applications"

        headers = ["Job Title", "Company Name", "Application Status", "Apply Date"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header

        row_num = 2
        for application in applications_cursor:
            job = db.jobs.find_one(
                {"_id": application['job_id']},
                {"title": 1, "company_name": 1}
            )

            if job:
                job_title = job.get("title", "")
                company_name = job.get("company_name", "")
                application_status = application.get("application_status", "")
                apply_date = application.get("created_at", datetime.now()).strftime("%Y-%m-%d %H:%M:%S")

                ws[f"A{row_num}"] = job_title
                ws[f"B{row_num}"] = company_name
                ws[f"C{row_num}"] = application_status
                ws[f"D{row_num}"] = apply_date

                row_num += 1

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=job_applications_report.xlsx"})

    except Exception as e:
        api_error(500, f"Error generating report: {str(e)}")


@router.get('/my-jobs/export')
def export_applications(user=Depends(require_role('employer'))):
        jobs_cursor = db.jobs.find({"post_by": ObjectId(user['_id'])})

        wb = Workbook(write_only=True)  
        ws = wb.create_sheet(title="My All Job")

        headers = ["Job Title", "Applications", "Status", "Posted", "Expires"]
        ws.append(headers)

        for job in jobs_cursor:
            job_title = job.get("title", "")
            status = job.get("job_status", "")
            posted = job.get("created_at", datetime.now()).strftime("%Y-%m-%d")
            expires = job.get("application_dead_line", datetime.now()).strftime("%Y-%m-%d")

            application_count = db.applications.count_documents({"job_id": job["_id"]})

            ws.append([
                job_title,
                application_count,
                status,
                posted,
                expires
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=my_jobs_export.xlsx"})


@router.get('/featured-jobs')
def featured_jobs():
    pipeline = [
        { "$sort": { "created_at": -1 } },
        { "$limit": 3 },
        {
            "$lookup": {
                "from": "companies",            
                "localField": "company_name",       
                "foreignField": "_id",    
                "as": "company"
            }
        },
        { "$unwind": "$company" },
        {
            "$project":{
                "_id":{"$toString": "$_id"},
                "company.name": 1,
                "job_type": 1,
                "body": 1,
                "location": 1,
                "title": 1,
                "skills": 1,
                "slug": 1
            }
        }
    ]

    jobs = list(db.jobs.aggregate(pipeline))

    return {
        "message": "success",
        "jobs": jobs
    }

@router.get('/my-jobs')
async def get_my_jobs(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1),
    user=Depends(require_role('employer'))
):
    skip = (page -1) * limit

    pipeline = [
        {"$match": {"post_by": ObjectId(user['_id'])}},
        {
                "$lookup": {
                    "from": "companies", 
                    "localField": "company_name",
                    "foreignField": "_id", 
                    "as": "posted_by_user"
                }
        },
        {"$unwind": {"path": "$posted_by_user", "preserveNullAndEmptyArrays": True}}, 
        {
        "$lookup": {
            "from": "applications",
            "localField": "_id",
            "foreignField": "job_id",
            "as": "applications"
        }
        },
        {
            "$addFields": {
                "total_applications": {"$size": "$applications"}
            }
        },
        {"$skip": skip},
        {"$limit": limit},
        {
            "$project":{
                 "_id": {"$toString": "$_id"}, 
                    "title": 1,
                    "job_type": 1,
                    "slug": 1,
                    "created_at": 1,
                    "job_status": 1,
                    "updated_at": 1,
                    "total_applications": 1,
                    "posted_by_user._id": {"$toString": "$posted_by_user._id"},
                    "posted_by_user.name": 1,
                    "posted_by_user.logo": 1
            }
        }
    ]

    jobs = list(db.jobs.aggregate(pipeline))


    return {
        "message": "success",
        "jobs": jobs
    }

@router.get('/employer/stats')
def get_employer_stats(user=Depends(require_role('employer'))):

    company_id = ObjectId(user['company_id'])

    jobs = list(db.jobs.find({"company_name": company_id}, {"_id": 1, "job_view": 1}))
    job_ids = [job["_id"] for job in jobs]

    active_job = db.jobs.count_documents({
        "company_name": company_id,
        "job_status": "active"
    })

    total_application = db.applications.count_documents({
        "job_id": {"$in": job_ids}
    })

    interviewed = db.applications.count_documents({
        "job_id": {"$in": job_ids},
        "application_status": "Interviewed"
    })

    jobs_view = sum(job.get("job_view", 0) for job in jobs)

    return {
        "active_jobs": active_job,
        "total_application": total_application,
        "interviewed": interviewed,
        "jobs_view": jobs_view
    }

@router.get("/employer/recent-applicants")
def get_recent_applicants(
    user=Depends(require_role("employer")),
    days: int = Query(10, ge=1),
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1),
):
    company_id = ObjectId(user["company_id"])
    since_date = datetime.now(timezone.utc) - timedelta(days=days)
    skip = (page - 1) * limit

    job_query = {"company_name": company_id, "job_status": "active"}

    jobs = list(db.jobs.find(job_query))
    job_id_map = {job["_id"]: job for job in jobs}
    job_ids = list(job_id_map.keys())

    if not job_ids:
        return {"message": "No jobs found", "applicants": [], "total": 0}

    pipeline = [
        {
            "$match": {
                "job_id": {"$in": job_ids},
                "created_at": {"$gte": since_date}
            }
        },
        {"$sort": {"created_at": -1}},
        {"$skip": skip},
        {"$limit": limit},
        {
            "$lookup": {
                "from": "users",
                "localField": "applicate",
                "foreignField": "_id",
                "as": "user"
            }
        },
        {"$unwind": {"path": "$user", "preserveNullAndEmptyArrays": True}},
        {
            "$project": {
                "_id": {"$toString": "$_id"},
                "job_id": 1,
                "application_status": 1,
                "created_at": 1,
                "user": 1
            }
        }
    ]

    apps = list(db.applications.aggregate(pipeline))
    total = db.applications.count_documents({
        "job_id": {"$in": job_ids},
        "created_at": {"$gte": since_date}
    })

    enriched = []
    for app in apps:
        job = job_id_map.get(app["job_id"])
        user_info = app.get("user", {})
        enriched.append({
            "_id": app["_id"],
            "job_id": str(app["job_id"]),
            "job_title": job.get("title", "Unknown Job") if job else "Unknown",
            "created_at": app["created_at"].isoformat(),
            "application_status": app.get("application_status"),
            "first_name": user_info.get("first_name", ""),
            "user_id": str(user_info.get("_id", "")),
            "last_name": user_info.get("last_name", ""),
            "applicant_email": user_info.get("email", ""),
            "resume": user_info.get("resume", ""),
            "match_percentage": calculate_match_percentage(job or {}, user_info)
        })

    return {
        "message": "success",
        "page": page,
        "limit": limit,
        "total": total,
        "applicants": enriched
    }





@router.get("/employer/active-jobs")
def get_active_jobs_with_application_count(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1),
    user=Depends(require_role("employer"))
):
    skip = (page - 1) * limit
    company_id = ObjectId(user["company_id"])

    jobs = list(db.jobs.find(
        {"company_name": company_id, "job_status": "active"},
        {
            "_id": 1,
            "title": 1,
            "location": 1,
            "job_type": 1,
            "created_at": 1,
            "application_dead_line": 1,
            "slug": 1
        }
    ).skip(skip).limit(limit))

    job_ids = [job["_id"] for job in jobs]
    application_counts = db.applications.aggregate([
        {"$match": {"job_id": {"$in": job_ids}}},
        {"$group": {"_id": "$job_id", "count": {"$sum": 1}}}
    ])
    count_map = {item["_id"]: item["count"] for item in application_counts}

    for job in jobs:
        job["total_applications"] = count_map.get(job["_id"], 0)
        job["_id"] = str(job["_id"])  

    return {"jobs": jobs}


@router.get('/companies')
def get_companies(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1),
    title: str = Query(None),
    location: str = Query(None),
    industry: str = Query(None)
):
    skip = (page-1) * limit
    query = {}
    if title:
            query["name"] = {"$regex": title, "$options": "i"}  
    if location:
            query["headquatar_location"] = {"$regex": location, "$options": "i"}
    if industry:
            query["industry"] = {"$regex": industry, "$options": "i"}

    pipeline = [
            {"$match": query},
            {"$sort": {"_id": -1}},  
            {"$skip": skip},
            {"$limit": limit},
            {
                "$lookup": {
                    "from": "jobs",
                    "let": { "companyId": "$_id" },
                    "pipeline": [
                        {
                            "$match": {
                                "$expr": {
                                    "$and": [
                                        { "$eq": ["$company_name", "$$companyId"] },
                                        { "$eq": ["$job_status", "active"] }
                                    ]
                                }
                            }
                        },
                        { "$count": "count" }
                    ],
                    "as": "active_jobs"
                }
            },
            {
                "$addFields": {
                    "jobs_open": {
                        "$ifNull": [{ "$arrayElemAt": ["$active_jobs.count", 0] }, 0]
                    }
                }
            },
            {
                "$project": {
                    "_id": { "$toString": "$_id" },
                    "name": 1,
                    "company_size": 1,
                    "logo": 1,
                    "company_description": 1,
                    "industry": 1,
                    "headquatar_location": 1,
                    "jobs_open": 1,
                    "rating": 1,
                }
            }
        ]

    companies = list(db.companies.aggregate(pipeline))
    total = db.companies.count_documents({})

    return {
            "message": "success",
            "page": page,
            "limit": limit,
            "total": total,
            "companies": companies
    }


@router.get('/company/{id}')
def get_single_company(id:str):

    pipeline = [
            {"$match": {"_id": ObjectId(id)}},
            {
                "$lookup": {
                    "from": "jobs",
                    "let": {"companyId": "$_id"},
                    "pipeline": [
                        {
                            "$match": {
                                "$expr": {
                                    "$and": [
                                        {"$eq": ["$company_name", "$$companyId"]},
                                        {"$eq": ["$status", "active"]}
                                    ]
                                }
                            }
                        },
                        {
                            "$project": {
                                "_id": 0,
                                "job_id": {"$toString": "$_id"},
                                "title": 1,
                                "slug": 1,
                                "job_type": 1,
                                "location": 1,
                                "min_salary": 1,
                                "max_salary": 1,
                                "created_at": 1
                            }
                        }
                    ],
                    "as": "active_jobs"
                }
            },
            {
                "$project": {
                    "_id": {"$toString": "$_id"},
                    "name": 1,
                    "logo": 1,
                    "rating": 1,
                    "website": 1,
                    "company_size": 1,
                    "company_description": 1,
                    "industry": 1,
                    "headquatar_location": 1,
                    "comapnies_calture": 1,
                    "active_jobs": 1,
                    "social_media": 1,
                    "ratings": {
                    "$map": {
                        "input": "$ratings",
                        "as": "r",
                        "in": {
                            "rating": "$$r.rating",
                            "employe_status": "$$r.employe_status",
                            "title": "$$r.title",
                            "review": "$$r.review",
                            "props": "$$r.props",
                            "cons": "$$r.cons",
                            "is_recommend": "$$r.is_recommend",
                            "date": "$$r.date",
                            "write_by": {"$toString": "$$r.write_by"}
                        }
                    }
        }
                }
            }
        ]

    companies = list(db.companies.aggregate(pipeline))

    if not companies:
            api_error(404, "Company not found")

    return {
            "message": "success",
            "company": companies[0]
        }

@router.post('/review/{id}')
def write_review(
    id: str,
    body: ReviewsRatinsSchema,
    user=Depends(require_role('employer'))
):
    company_id = ObjectId(id)
    reviewer_id = ObjectId(user['_id'])

    existing_review = db.companies.find_one({
        "_id": company_id,
        "ratings.write_by": reviewer_id
    })

    if existing_review:
        api_error(400, "You have already submitted a review for this company.")

    payload = {
        **body.model_dump(),
        "write_by": reviewer_id,
        "date": datetime.now(timezone.utc)
    }

    db.companies.update_one(
        {"_id": company_id},
        {"$push": {"ratings": payload}}
    )

    company = db.companies.find_one({"_id": company_id})
    all_ratings = company.get("ratings", [])

    if all_ratings:
        average = sum(r.get("rating", 0) for r in all_ratings if r.get("rating") is not None) / len(all_ratings)
        db.companies.update_one(
            {"_id": company_id},
            {"$set": {"rating": round(average, 2)}}
        )

    return {
        "message": "Review submitted successfully."
    }

@router.get('/candidates')
def get_candidates(
    user=Depends(require_role('employer')),
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1)
):

    skip = (page-1) * limit
    jobs = list(db.jobs.find({"post_by": ObjectId(user["_id"])}, {"_id": 1}))
    job_ids = [job["_id"] for job in jobs]

    if not job_ids:
            return {
                "message": "No jobs posted by this employer",
                "applicants": []
            }

    pipeline = [
            {
                "$match": {
                    "job_id": {"$in": job_ids}
                }
            },
            {"$skip": skip},
            {"$limit": limit},
            {
                "$lookup": {
                    "from": "users",
                    "localField": "applicate",
                    "foreignField": "_id",
                    "as": "applicant"
                }
            },
            {"$unwind": "$applicant"},
            {
                "$lookup": {
                    "from": "jobs",
                    "localField": "job_id",
                    "foreignField": "_id",
                    "as": "job"
                }
            },
            {"$unwind": "$job"},
            {
                "$project": {
                    "_id": {"$toString": "$_id"},
                    "job_id": {"$toString": "$job._id"},
                    "job_title": "$job.title",
                    "first_name": "$applicant.first_name",
                    "avatar": "$applicant.avatar",
                    "last_name": "$applicant.last_name",
                    "email": "$applicant.email",
                    "resume": "$applicant.resume",
                    "location": "$applicant.location",
                    "application_status": 1,
                    "cover_letter": 1,
                    "experience": 1,
                    "skills": "$job.skills",
                    "applied_at": "$created_at"
                }
            }
        ]

    applicants = list(db.applications.aggregate(pipeline))

    return {
            "message": "success",
            "total": len(applicants),
            "applicants": applicants
        }

@router.get("/job-apps/{user_id}")
async def get_spefic_job_applications(
    user_id: str, 
    user=Depends(require_role('employer'))
):
    pipeline = [
        {"$match": {"applicate": ObjectId(user_id)}},
        {
            "$lookup": {
                "from": "users",
                "localField": "applicate",
                "foreignField": "_id",
                "as": "applicant"
            }
        },
        {"$unwind": "$applicant"},
        {
            "$lookup": {
                "from": "jobs",
                "localField": "job_id",
                "foreignField": "_id",
                "as": "job"
            }
        },
        {"$unwind": "$job"},
        {
            "$project": {
                "_id": {"$toString": "$_id"},
                "user_id": {"$toString": "$applicant._id"},
                "job_id": {"$toString": "$job._id"},
                "job_title": "$job.title",
                "first_name": "$applicant.first_name",
                "avatar": "$applicant.avatar",
                "last_name": "$applicant.last_name",
                "email": "$applicant.email",
                "resume": "$applicant.resume",
                "location": "$applicant.location",
                "education": "$applicant.education",
                "user_experience": "$applicant.experience",
                "github_profile": "$applicant.github_profile",
                "linkedin_profile": "$applicant.linkedin_profile",
                "website": "$applicant.website",
                "phone_number": "$applicant.phone_number",
                "application_status": 1,
                "cover_letter": 1,
                "experience": 1,
                "questions_answer": 1,
                "summary": "$applicant.summary",
                "applied_at": "$created_at",
                "job_skills": "$job.skills",
                "user_skills": "$applicant.skills"
            }
        }
    ]

    raw_data = list(db.applications.aggregate(pipeline))

    if not raw_data:
        return {"message": "No application found", "applicants": None}

    for app in raw_data:
        job_info = {"skills": app.get("job_skills", [])}
        user_info = {"skills": app.get("user_skills", [])}
        app["match_percentage"] = calculate_match_percentage(job_info, user_info)
        resume_path = app.get("resume")
        if resume_path:
            try:
                text = extract_text_from_pdf(resume_path)
                app["resume_summary"] = extract_structured_data_from_resume(text)
            except Exception as e:
                app["resume_summary"] = ""
        else:
            app["resume_summary"] = ""

    db.users.find_one_and_update(
        {"_id": ObjectId(user_id)},
        {"$inc": {"profile_view": 1}}
    )
    await emit_event(user_id, "PROFILE_VIEW", {"view": 1})
    return {
        "message": "Success",
        "applicants": raw_data[0] 
    }

@router.post("/generate-cover-letter")
def generate_cover_letter(data:CoverLetterInput, user=Depends(require_role('job_seeker'))):

    prompt = f"""
    Write a professional cover letter for the following candidate:

    Name: {data.full_name}
    Position: {data.position}
    Company: {data.company_name}
    Years of Experience: {data.experience}
    Skills: {", ".join(data.skills)}

    The tone should be formal and confident.
    """

    response = client.chat.completions.create(
            model=AI_MODEL,  
            messages=[{"role": "user", "content": prompt}],
            temperature=0.2
        )
    
    generated_text = response.choices[0].message.content.strip() or 'not able to generate cover letter '
    return { "cover_letter": generated_text }


@router.get("/get-applications/{job_id}")
def get_job_spefic_applications(
    job_id:str, 
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1),
    user=Depends(require_role('employer'))
):
    job = db.jobs.find_one({"_id": ObjectId(job_id)})
    if not job:
        api_error(404, "Job not found")

    skip = (page-1) * limit
    pipeline = [
        {"$match": {
            "job_id": ObjectId(job_id)
        }},
        {
            "$lookup":{
                "from": "users",
                "localField":"applicate",
                "foreignField":"_id",
                "as": "user"
            }
        },
        {"$unwind": {"path": "$user", "preserveNullAndEmptyArrays": True}},
        {
            "$lookup":{
                "from": "jobs",
                "localField":"job_id",
                "foreignField":"_id",
                "as": "job"
            }
        },
        {"$unwind": {"path": "$job", "preserveNullAndEmptyArrays": True}},
        {"$skip": skip},
        {"$limit": limit},
        {
            "$project":{
                "_id": {"$toString": "$_id"},
                "job_id": {"$toString": "$job._id"},
                "user_id": {"$toString": "$user._id"},
                "title": "$job.title",
                "email": "$user.email",
                "avatar": "$user.avatar",
                "first_name": "$user.first_name",
                "last_name": "$user.last_name",
                "application_status":1,
                "created_at": 1,
                "job_skills": "$job.skills",
                "user_skills": "$user.skills",
                "resume": "$user.resume",
            }
        }
    ]

    raw_data = list(db.applications.aggregate(pipeline))

    for app in raw_data:
        job_info = {"skills": app.get("job_skills", [])}
        user_info = {"skills": app.get("user_skills", [])}
        app["match_percentage"] = calculate_match_percentage(job_info, user_info)
    
    total = db.applications.count_documents({ "job_id": ObjectId(job_id)})
    
    return {
        "message":"success",
        "title": job.get("title", ""),
        "applications": raw_data,
        "total": total
    }